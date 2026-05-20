#!/usr/bin/env python3
"""
Import excavation camp participation from XLS files into pvh_avm_camp_participation.

Usage:
    python3 import-avpvh-camps.py [--dry-run]

Iterates over all .xls/.xlsx files under /home/grmt/avpvh_drive/xls/Opgravingen/.
For each file:
  - Extracts year and location from directory name / filename
  - Creates pvh_avm_camps row if not yet present
  - Reads "totaal inschrijvingen" sheet (or first sheet) for participants
  - Matches participants to pvh_avm_members by last_name + first_name
  - Logs unmatched names to stdout for manual review
"""

import argparse
import os
import re
import sys
from pathlib import Path

import pymysql
import openpyxl

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

SECRET_FILE   = '/opt/docker/secrets/compose/mysql_password.txt'
DB_HOST       = '127.0.0.1'
DB_PORT       = 6603
DB_USER       = 'wordpress'
DB_NAME       = 'wordpress_pvh'
WP_PREFIX     = 'pvh_'
OPRAVINGEN_ROOT = Path('/home/grmt/avpvh_drive/xls/Opgravingen')


def read_db_password() -> str:
    with open(SECRET_FILE) as f:
        return f.read().strip()


def get_connection(password: str):
    return pymysql.connect(
        host=DB_HOST, port=DB_PORT,
        user=DB_USER, password=password,
        database=DB_NAME, charset='utf8mb4',
    )


def extract_year_location(path: Path) -> tuple[int | None, str]:
    # Try to get year from directory name first, then filename
    year = None
    for part in reversed(path.parts):
        m = re.search(r'\b(20\d{2}|19\d{2})\b', part)
        if m:
            year = int(m.group())
            break

    # Location from parent directory name (strip year)
    location = re.sub(r'\b(20|19)\d{2}\b', '', path.parent.name).strip(' _-')
    if not location:
        location = re.sub(r'\b(20|19)\d{2}\b', '', path.stem).strip(' _-')

    return year, location


def open_workbook(path: Path):
    suffix = path.suffix.lower()
    if suffix == '.xlsx':
        return openpyxl.load_workbook(path, data_only=True), 'openpyxl'
    elif suffix == '.xls' and HAS_XLRD:
        return xlrd.open_workbook(str(path)), 'xlrd'
    return None, None


def get_sheet(wb, wb_type: str, preferred_name: str):
    if wb_type == 'openpyxl':
        names_lower = {s.lower(): s for s in wb.sheetnames}
        key = preferred_name.lower()
        if key in names_lower:
            return wb[names_lower[key]], 'openpyxl'
        return wb.active, 'openpyxl'
    elif wb_type == 'xlrd':
        for i in range(wb.nsheets):
            if preferred_name.lower() in wb.sheet_by_index(i).name.lower():
                return wb.sheet_by_index(i), 'xlrd'
        return wb.sheet_by_index(0), 'xlrd'
    return None, None


def iter_rows_as_strings(sheet, sheet_type: str):
    if sheet_type == 'openpyxl':
        for row in sheet.iter_rows(values_only=True):
            yield [str(c).strip() if c is not None else '' for c in row]
    elif sheet_type == 'xlrd':
        for rx in range(sheet.nrows):
            yield [str(sheet.cell_value(rx, cx)).strip() for cx in range(sheet.ncols)]


def find_member(cursor, last_name: str, first_name: str) -> int | None:
    cursor.execute(
        f"""SELECT id FROM {WP_PREFIX}avm_members
            WHERE LOWER(last_name) = LOWER(%s) AND LOWER(first_name) = LOWER(%s)
            LIMIT 1""",
        (last_name, first_name)
    )
    row = cursor.fetchone()
    return row[0] if row else None


def fuzzy_find_member(cursor, full_name: str) -> int | None:
    parts = full_name.strip().split()
    if len(parts) < 2:
        return None
    # Try last word as last name, rest as first
    last  = parts[-1]
    first = ' '.join(parts[:-1])
    mid = find_member(cursor, last, first)
    if mid:
        return mid
    # Try first word as first name
    first2 = parts[0]
    last2  = ' '.join(parts[1:])
    return find_member(cursor, last2, first2)


def import_file(cursor, path: Path, dry_run: bool, unmatched: list):
    year, location = extract_year_location(path)
    if not year:
        print(f'  SKIP (no year found): {path}')
        return

    camp_name = path.parent.name
    print(f'\n{path.name}  →  camp="{camp_name}" year={year} location="{location}"')

    # Ensure camp row
    cursor.execute(
        f'SELECT id FROM {WP_PREFIX}avm_camps WHERE name = %s AND year = %s',
        (camp_name, year)
    )
    camp_row = cursor.fetchone()
    if camp_row:
        camp_id = camp_row[0]
    elif not dry_run:
        cursor.execute(
            f'INSERT INTO {WP_PREFIX}avm_camps (name, year, location) VALUES (%s,%s,%s)',
            (camp_name, year, location)
        )
        camp_id = cursor.lastrowid
        print(f'  created camp id={camp_id}')
    else:
        camp_id = -1

    wb, wb_type = open_workbook(path)
    if not wb:
        print(f'  SKIP (unreadable format): {path}')
        return

    sheet, sheet_type = get_sheet(wb, wb_type, 'totaal inschrijvingen')
    rows = list(iter_rows_as_strings(sheet, sheet_type))
    if not rows:
        print('  SKIP (empty sheet)')
        return

    # Detect header row
    header_idx = 0
    headers = []
    for i, row in enumerate(rows[:10]):
        row_lower = [c.lower() for c in row]
        if any(k in row_lower for k in ('naam', 'achternaam', 'voornaam')):
            header_idx = i
            headers = row_lower
            break

    if not headers:
        print('  SKIP (no recognisable header row)')
        return

    def col_idx(name: str) -> int | None:
        try:
            return headers.index(name)
        except ValueError:
            return None

    idx_last   = col_idx('achternaam') or col_idx('naam')
    idx_first  = col_idx('voornaam')
    idx_nights = col_idx('nachten') or col_idx('nights')
    idx_nawacht= col_idx('nawacht')
    idx_diet   = col_idx('dieet') or col_idx('diet')
    idx_notes  = col_idx('notities') or col_idx('opmerkingen') or col_idx('notes')

    imported = skipped = 0
    for row in rows[header_idx + 1:]:
        if idx_last is not None and idx_last < len(row):
            last_name = row[idx_last]
        else:
            continue
        if not last_name or last_name.lower() in ('totaal', 'sum', ''):
            continue

        first_name = row[idx_first] if idx_first is not None and idx_first < len(row) else ''

        member_id = find_member(cursor, last_name, first_name)
        if not member_id and first_name:
            member_id = fuzzy_find_member(cursor, f'{first_name} {last_name}')
        if not member_id:
            full = f'{last_name}, {first_name}'.strip(', ')
            unmatched.append({'file': str(path), 'name': full})
            print(f'  UNMATCHED: {full}')
            skipped += 1
            continue

        nights  = _parse_int(row[idx_nights])  if idx_nights  and idx_nights  < len(row) else None
        nawacht = _parse_int(row[idx_nawacht]) if idx_nawacht and idx_nawacht < len(row) else 0
        diet    = row[idx_diet][:50]           if idx_diet    and idx_diet    < len(row) else None
        notes   = row[idx_notes]               if idx_notes   and idx_notes   < len(row) else None

        if not dry_run and camp_id > 0:
            cursor.execute(
                f"""INSERT IGNORE INTO {WP_PREFIX}avm_camp_participation
                    (member_id, camp_id, nights, nawacht, diet, notes)
                    VALUES (%s,%s,%s,%s,%s,%s)""",
                (member_id, camp_id, nights, nawacht or 0, diet or None, notes or None)
            )
        imported += 1

    print(f'  imported={imported} skipped(unmatched)={skipped}')


def _parse_int(val: str) -> int | None:
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()

    if not OPRAVINGEN_ROOT.exists():
        sys.exit(f'ERROR: {OPRAVINGEN_ROOT} does not exist')

    password = read_db_password()
    conn = get_connection(password)
    unmatched = []

    try:
        with conn.cursor() as cur:
            xls_files = sorted(OPRAVINGEN_ROOT.rglob('*.xls')) + sorted(OPRAVINGEN_ROOT.rglob('*.xlsx'))
            print(f'Found {len(xls_files)} XLS/XLSX files under {OPRAVINGEN_ROOT}')
            for path in xls_files:
                import_file(cur, path, args.dry_run, unmatched)

        if not args.dry_run:
            conn.commit()
            print('\nCommitted.')
        else:
            print('\nDry-run complete — no changes written.')
    finally:
        conn.close()

    if unmatched:
        print(f'\n=== {len(unmatched)} UNMATCHED PARTICIPANTS (manual review needed) ===')
        for u in unmatched:
            print(f'  {u["name"]}  ({Path(u["file"]).name})')


if __name__ == '__main__':
    main()
