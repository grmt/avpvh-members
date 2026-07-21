#!/usr/bin/env python3
"""
Sync the excavation campaign's day-by-day participation grid (sheet "totaal
inschrijvingen" in the club's registration spreadsheet) into a WordPress
option as JSON, for the [avpvh_kamp_overzicht] shortcode to render.

This is a snapshot, not a live view — the spreadsheet changes continuously
as people register/cancel, so re-run this whenever the site should show
fresher data. No LLDAP/member-account writes here, just one wp_options row.

Usage:
    python3 sync-kamp-overzicht.py /path/to/Overzicht.xlsx [--dry-run]

Dependencies:
    pip install pymysql openpyxl
"""
import argparse
import json

import pymysql
import openpyxl

SECRET_FILE = '/opt/docker/secrets/compose/wordpress_db_password.txt'
DB_HOST, DB_PORT, DB_USER, DB_NAME = '127.0.0.1', 6603, 'wp_user', 'wpdb'
WP_PREFIX = 'pvh_'
OPTION_NAME = 'avpvh_kamp_2026_overzicht'

SHEET_NAME = 'totaal inschrijvingen'
GRID_FIRST_ROW = 3
GRID_LAST_ROW = 97
GRID_FIRST_COL = 4   # D
GRID_LAST_COL = 26   # Z


def read_secret(path: str) -> str:
    with open(path) as f:
        return f.read().strip()


def cell_color(cell) -> str | None:
    fill = cell.fill
    if not fill or not fill.fgColor:
        return None
    rgb = fill.fgColor.rgb
    if not rgb or not isinstance(rgb, str) or rgb in ('00000000',):
        return None
    # openpyxl ARGB -> #RRGGBB
    return '#' + rgb[-6:]


def cell_bold(cell) -> bool:
    return bool(cell.font and cell.font.bold)


def cell_value(cell) -> str:
    v = cell.value
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def build_grid(ws) -> dict:
    title = cell_value(ws.cell(row=1, column=3)) or 'Overzicht inschrijvingen'
    last_updated = cell_value(ws.cell(row=2, column=3))
    note = cell_value(ws.cell(row=2, column=6))

    grid = []
    for r in range(GRID_FIRST_ROW, GRID_LAST_ROW + 1):
        row = []
        for c in range(GRID_FIRST_COL, GRID_LAST_COL + 1):
            cell = ws.cell(row=r, column=c)
            row.append({'v': cell_value(cell), 'c': cell_color(cell), 'b': cell_bold(cell)})
        grid.append(row)

    return {'title': title, 'last_updated': last_updated, 'note': note, 'grid': grid}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx', help='Path to the Overzicht Inschrijvingen xlsx')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f'sheet {SHEET_NAME!r} not found; sheets: {wb.sheetnames}')

    data = build_grid(wb[SHEET_NAME])
    payload = json.dumps(data, ensure_ascii=False)
    print(f'Title: {data["title"]!r}')
    print(f'Last updated (per sheet): {data["last_updated"]!r}')
    print(f'Grid: {len(data["grid"])} rows x {GRID_LAST_COL - GRID_FIRST_COL + 1} cols')
    print(f'Payload size: {len(payload)} bytes')

    if args.dry_run:
        print('\nDry-run complete — option not written.')
        return

    conn = pymysql.connect(host=DB_HOST, port=DB_PORT, user=DB_USER,
                            password=read_secret(SECRET_FILE), database=DB_NAME,
                            charset='utf8mb4')
    try:
        with conn.cursor() as cur:
            cur.execute(
                f"INSERT INTO {WP_PREFIX}options (option_name, option_value, autoload) "
                f"VALUES (%s, %s, 'no') "
                f"ON DUPLICATE KEY UPDATE option_value = VALUES(option_value)",
                (OPTION_NAME, payload)
            )
        conn.commit()
        print(f'\nWritten to wp_options[{OPTION_NAME}].')
    finally:
        conn.close()


if __name__ == '__main__':
    main()
