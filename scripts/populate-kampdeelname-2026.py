#!/usr/bin/env python3
"""
One-off population of avm_camp_participation(_days) for the 2026 Goeblange
camp, from the "totaal inschrijvingen" sheet in the club's registration
spreadsheet — the structured replacement for the old raw-grid snapshot
used by [avpvh_kamp_overzicht] (see class-kamp-overzicht.php / sync-kamp-
overzicht.py, which only ever captured display formatting, not semantics).

The name -> member_id mapping below was derived by cross-matching the
sheet's 75 signed-up names against the live avm_members list (fuzzy
first+last name matching, then manually confirmed for spelling variants
and nicknames). People not yet in avm_members (mostly relatives/partners
of existing members who haven't got their own member record) are listed
in UNLINKED_NAMES and are skipped with a warning — add them as members
first, then re-run.

Usage:
    python3 populate-kampdeelname-2026.py "/path/to/Overzicht Inschrijvingen Opgraving 2026 Goeblange.xlsx" [--dry-run]

Dependencies:
    pip install pymysql openpyxl
"""
import argparse
import datetime
import re

import pymysql
import openpyxl


def normalize_name(name: str) -> str:
    """Collapse whitespace and unify hyphenated-surname spelling. The sheet
    writes double surnames as "Hoek - De Boe" (spaced, plain hyphen); the
    actual avm_members record has no spaces around the hyphen ("Hoek-De
    Boe") — and the browser-scraped snapshot used to build
    NAME_TO_MEMBER_ID had yet another variant (en-dash, spaced). Unify all
    of these to "Hoek-De Boe" so the lookup isn't sensitive to any of it."""
    name = re.sub(r'[‐-―]', '-', name)
    name = re.sub(r'\s*-\s*', '-', name)
    return re.sub(r'\s+', ' ', name).strip()

SECRET_FILE = '/opt/docker/secrets/compose/wordpress_db_password.txt'
DB_HOST, DB_PORT, DB_USER, DB_NAME = '127.0.0.1', 6603, 'wp_user', 'wpdb'
WP_PREFIX = 'pvh_'

SHEET_NAME = 'totaal inschrijvingen'
GRID_FIRST_ROW = 3
GRID_LAST_ROW = 97
GRID_FIRST_COL = 4   # D: name
# Column offsets within a row, relative to GRID_FIRST_COL (0-based)
COL_NAME = 0
COL_DAYS = range(1, 17)     # E..U: 2026-07-24 .. 2026-08-08 (16 days)
COL_NIGHTS = 17              # V: Aantal Nachten
COL_TOTAL = 18                # W: Totaal per Persoon (financial, not imported)
COL_NAWACHT = 19              # X: Nawacht ("ja" / blank)
COL_OPMERKINGEN = 20          # Y: aankomst/vertrek + bijzonderheden
COL_OVERIGE = 21               # Z... (see note below)
COL_DIET = 22                   # last column: Dieet

CAMP_NAME = 'Goeblange'
CAMP_YEAR = 2026
CAMP_START = datetime.date(2026, 7, 24)
CAMP_END = datetime.date(2026, 8, 8)

# Section/summary rows to always skip (not real people).
SKIP_LABELS = {
    '', 'Volwassenen', 'Tieners', 'Kinders', 'Peuters', 'Aantal overnachters',
    'VOORNAAM & ACHTERNAAM hieronder invullen', 'Bestuur', 'Leden',
}

# People on the sheet who are NOT yet avm_members (as of 2026-07-23) — run
# add-2026-visitor-members.py first (creates them as status='visitor'), then
# move each name from UNLINKED_NAMES into NAME_TO_MEMBER_ID below using the
# member ids it prints, and re-run this script.
UNLINKED_NAMES = {
    'Fenna Lip',                    # daughter of Henk & Mariska Lip, sister of Dirk(118)/Roos(121)
    'Iris de Zwart',                # partner of Olaf Boekholt (32)
    'Jessica Hammarlund Bergmann',
    'May Hasendonckx',              # partner of Gerrit Hasendonckx (51)
    'Taras Muravskiy',
    'Dean Berendsen',               # son of Sylvia Soulier (80)
    'Bram Keijers',                 # relative of Jaap Keijers (57) — not the same person
}

# name (exactly as it appears in the sheet) -> avm_members.id
NAME_TO_MEMBER_ID = {
    'Annet van de Waarsenburg': 88, 'Ardine de Wit': 89, 'Germie van den Berg': 29,
    'José Salhi-Vossen': 72, 'Lilo Crasborn': 40, 'Meike Waaijers': 87,
    'Rianne de Wit': 91, 'Simon Kuppens': 58, 'Aukje Sinnema': 75,
    'Barbara Hoek-De Boe': 105, 'Bas Waaijers': 85, 'Bernard te Gussinklo': 117,
    'Claudia Boshouwers': 34, 'Dirk Lip': 118, 'Eefje Smits': 77,
    'Elout Crasborn': 39, 'Elvire Rikkert': 106, 'Femke Hasendonckx': 50,
    'Frank Clermonts': 37, 'Frank Schoots': 73, 'Frank van Rooij': 67,
    'Garmt Boekholt': 31, 'Gaston Moonen': 65, 'Geert Smits': 107,
    'Gerrit Hasendonckx': 51, 'Hélène Rikkert': 108, 'Jac. Biemans': 30,
    'Jessica Waaijers': 86, 'Jop Sutmuller': 81, 'Joris Clermonts': 38,
    'Juba Salhi': 71, 'Jules van Horen': 56, 'Juul Vervuren': 83,
    'Katja Soulier': 79, 'Kitty de Bruin': 36, 'Lisette van Erp': 46,
    'Mariska Hoogendijk': 54, 'Mathijs Borms': 33, 'Meereke Bosua': 35,
    'Mike Rietveld': 102, 'Naomi te Gussinklo': 48, 'Nele Sinnema': 76,
    'Nik Bloemers': 119, 'Novia de Wit': 90, 'Olaf Boekholt': 32,
    'Olivia te Gussinklo': 42, 'Paul Frissen': 47, 'Paula van den Kerkhof': 98,
    'Peter Schroeten': 74, 'Pim Vervuren': 103, 'Rebecca Luijten': 64,
    'Reijer Crasborn': 43, 'Roos Lip': 121, 'Sandra Driessen': 45,
    'Sam Leurink': 122, 'Saskia Smits': 78, 'Steef Boekholt': 123,
    'Sterre Vervuren': 110, 'Sylvia Soulier': 80, 'Thijmen Rikkert': 111,
    'Twan van Rooij': 69, 'Werner van Hoof': 53, 'Wouter Vienne': 113,
    'Michiel Hoek': 52, 'Pieter Hoek': 101, 'Roos Vienne': 99,
    'Doutzen Vienne': 100, 'Jaap Keijers': 57,
}


def read_secret(path: str) -> str:
    with open(path) as f:
        return f.read().strip()


def cell_value(cell) -> str:
    v = cell.value
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def parse_rows(ws):
    rows = []
    for r in range(GRID_FIRST_ROW, GRID_LAST_ROW + 1):
        row = [cell_value(ws.cell(row=r, column=GRID_FIRST_COL + c)) for c in range(23)]
        rows.append(row)
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx', help='Path to the Overzicht Inschrijvingen xlsx')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f'sheet {SHEET_NAME!r} not found; sheets: {wb.sheetnames}')
    rows = parse_rows(wb[SHEET_NAME])

    unlinked_norm = {normalize_name(n) for n in UNLINKED_NAMES}
    name_to_id_norm = {normalize_name(n): mid for n, mid in NAME_TO_MEMBER_ID.items()}

    person_rows = []
    for row in rows:
        name = normalize_name(row[COL_NAME])
        if name in SKIP_LABELS or name == '':
            continue
        if all(c == '' for c in row[1:]):
            continue
        person_rows.append(row)

    print(f'{len(person_rows)} people found on the sheet.')

    to_import = []
    for row in person_rows:
        name = normalize_name(row[COL_NAME])
        if name in unlinked_norm:
            print(f'  SKIP (not yet a member): {name}')
            continue
        member_id = name_to_id_norm.get(name)
        if not member_id:
            print(f'  SKIP (unrecognised name, check spelling / add to mapping): {name}')
            continue
        to_import.append((member_id, name, row))

    print(f'{len(to_import)} participation records will be written.')

    if args.dry_run:
        print('\nDry-run complete — nothing written.')
        return

    conn = pymysql.connect(host=DB_HOST, port=DB_PORT, user=DB_USER,
                            password=read_secret(SECRET_FILE), database=DB_NAME,
                            charset='utf8mb4')
    try:
        with conn.cursor() as cur:
            cur.execute(
                f"INSERT INTO {WP_PREFIX}avm_camps (name, year, location, start_date, end_date) "
                f"VALUES (%s, %s, %s, %s, %s) "
                f"ON DUPLICATE KEY UPDATE start_date = VALUES(start_date), end_date = VALUES(end_date)",
                (CAMP_NAME, CAMP_YEAR, 'Goeblange', CAMP_START, CAMP_END)
            )
            cur.execute(
                f"SELECT id FROM {WP_PREFIX}avm_camps WHERE name = %s AND year = %s",
                (CAMP_NAME, CAMP_YEAR)
            )
            camp_id = cur.fetchone()[0]

            for member_id, name, row in to_import:
                nights = row[COL_NIGHTS] or None
                nawacht = 1 if row[COL_NAWACHT].strip().lower() == 'ja' else 0
                diet = row[COL_DIET]
                notes_parts = [p for p in (row[COL_OPMERKINGEN], row[COL_OVERIGE]) if p]
                notes = ' / '.join(notes_parts)

                cur.execute(
                    f"INSERT INTO {WP_PREFIX}avm_camp_participation (member_id, camp_id, nights, nawacht, diet, notes) "
                    f"VALUES (%s, %s, %s, %s, %s, %s) "
                    f"ON DUPLICATE KEY UPDATE nights=VALUES(nights), nawacht=VALUES(nawacht), diet=VALUES(diet), notes=VALUES(notes)",
                    (member_id, camp_id, nights, nawacht, diet, notes)
                )
                cur.execute(
                    f"SELECT id FROM {WP_PREFIX}avm_camp_participation WHERE member_id = %s AND camp_id = %s",
                    (member_id, camp_id)
                )
                participation_id = cur.fetchone()[0]

                cur.execute(
                    f"DELETE FROM {WP_PREFIX}avm_camp_participation_days WHERE participation_id = %s",
                    (participation_id,)
                )
                for i, col in enumerate(COL_DAYS):
                    status = row[col]
                    if not status:
                        continue
                    date = CAMP_START + datetime.timedelta(days=i)
                    cur.execute(
                        f"INSERT INTO {WP_PREFIX}avm_camp_participation_days (participation_id, date, status) "
                        f"VALUES (%s, %s, %s)",
                        (participation_id, date, status)
                    )
        conn.commit()
        print(f'\nDone — {len(to_import)} participation records written for camp id {camp_id}.')
    finally:
        conn.close()


if __name__ == '__main__':
    main()
