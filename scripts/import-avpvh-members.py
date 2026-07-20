#!/usr/bin/env python3
"""
Import AVP-PvH members from the XLS ledenlijst into LLDAP + pvh_avm_* tables.

Usage:
    python3 import-avpvh-members.py /path/to/ledenlijst.xlsx [--dry-run]

Sheet "Leden"    → status = active  → LLDAP group "leden"
Sheet "ex-leden" → status = inactive → LLDAP group "ex-leden"

Idempotent on email. WP user creation is deferred to first login.

To update existing members' fields (birth date, address, phone, etc.) from a
newer spreadsheet, use reconcile-members.py instead — this script only
inserts new members and silently skips anyone whose derived LLDAP uid
already has a member record.

Dependencies:
    pip install pymysql openpyxl requests
"""

import argparse
from datetime import date

import openpyxl
import requests

from _avpvh_import_common import (
    WP_PREFIX, GROUP_ACTIVE, GROUP_INACTIVE,
    read_secret, get_db, lldap_login, get_group_id,
    uid_from_email, lldap_create_user, lldap_add_to_group,
    sheet_headers, col, parse_date, parse_year, age_on, placeholder_child_uid,
    SECRET_FILE,
)

CURRENT_YEAR = date.today().year

# ex-leden sheet has no header row; col 0 contains a left-year note
EX_LEDEN_COLS = ['vertrekjaar', 'achternaam', 'voornaam', 'geboortedatum',
                 'straat', 'huisnummer', 'postcode', 'woonplaats',
                 'land', 'email', 'telefoon', 'mobiel']


def import_sheet(cursor, session: requests.Session, sheet,
                 status: str, group_id: int, dry_run: bool,
                 positional_headers: list[str] | None = None) -> None:

    if positional_headers is not None:
        headers = positional_headers
        rows = sheet.iter_rows()
    else:
        headers = sheet_headers(sheet)
        rows = sheet.iter_rows(min_row=2)

    def c(row, name): return col(row, headers, name)

    for row in rows:
        email = c(row, 'email').removeprefix('mailto:')
        if not email or '@' not in email:
            continue

        last_name  = c(row, 'achternaam') or c(row, 'naam')
        suffix     = c(row, 'suffix')
        first_name = c(row, 'voornaam')
        phone      = c(row, 'telefoon')
        mobile     = c(row, 'mobiel')
        birth_date = parse_date(c(row, 'geboortedatum'))
        joined_yr  = parse_year(c(row, 'lidjaar') or c(row, 'lid jaar'))
        left_yr    = parse_year(c(row, 'vertrekjaar') or c(row, 'vertrek jaar'))
        emergency  = c(row, 'noodcontact') or c(row, 'emergency_contact')
        street     = c(row, 'straat')
        house_nr   = c(row, 'huisnummer') or c(row, 'nr')
        postal     = c(row, 'postcode')
        city       = c(row, 'woonplaats') or c(row, 'stad')
        country    = c(row, 'land') or 'Nederland'

        if birth_date and age_on(birth_date, date.today()) < 16:
            uid = placeholder_child_uid(session, first_name, last_name, dry_run)
            email = f'{uid}@avpvh.local'
        else:
            uid = uid_from_email(email)

        # Skip if this LLDAP uid already has a member record (family sharing one email)
        cursor.execute(
            f"SELECT id FROM {WP_PREFIX}avm_members WHERE lldap_user_id = %s",
            (uid,)
        )
        if cursor.fetchone():
            print(f'  skip (gezin deelt account): {last_name}, {first_name} → uid={uid}')
            continue

        # 1. Create LLDAP user
        lldap_create_user(session, uid, email, first_name, last_name, dry_run)
        lldap_add_to_group(session, uid, group_id, dry_run)

        if not dry_run:
            # 2. Insert pvh_avm_members
            cursor.execute(
                f"""INSERT INTO {WP_PREFIX}avm_members
                    (lldap_user_id, first_name, suffix, last_name, birth_date,
                     phone, mobile, emergency_contact, status, joined_year, left_year)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (uid, first_name, suffix, last_name, birth_date,
                 phone, mobile, emergency, status, joined_yr, left_yr)
            )
            member_id = cursor.lastrowid

            # 3. Insert address
            if street or city:
                cursor.execute(
                    f"""INSERT INTO {WP_PREFIX}avm_addresses
                        (member_id, street, house_number, postal_code, city, country, valid_until)
                        VALUES (%s,%s,%s,%s,%s,%s,NULL)""",
                    (member_id, street, house_nr, postal, city, country)
                )

            # 4. Insert pending fee for current year (active members only)
            if status == 'active':
                cursor.execute(
                    f"""INSERT IGNORE INTO {WP_PREFIX}avm_fees (member_id, year, status)
                        VALUES (%s,%s,'pending')""",
                    (member_id, CURRENT_YEAR)
                )

        print(f'  imported ({status}): {last_name}, {first_name} <{email}> → uid={uid}')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('xlsx', help='Path to ledenlijst xlsx')
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()

    db_password = read_secret(SECRET_FILE)
    conn = get_db(db_password)

    session = requests.Session()
    lldap_login(session)

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    print(f'Sheets: {wb.sheetnames}')

    try:
        with conn.cursor() as cur:
            group_active   = get_group_id(session, GROUP_ACTIVE)
            group_inactive = get_group_id(session, GROUP_INACTIVE)
            print(f'Groups: {GROUP_ACTIVE}={group_active}, {GROUP_INACTIVE}={group_inactive}')

            if 'Leden' in wb.sheetnames:
                print('\n=== Active members (sheet "Leden") ===')
                import_sheet(cur, session, wb['Leden'], 'active', group_active, args.dry_run)
            else:
                print('WARNING: sheet "Leden" not found')

            if 'ex-leden' in wb.sheetnames:
                print('\n=== Ex-members (sheet "ex-leden") ===')
                import_sheet(cur, session, wb['ex-leden'], 'inactive', group_inactive, args.dry_run,
                             positional_headers=EX_LEDEN_COLS)
            else:
                print('WARNING: sheet "ex-leden" not found')

        if not args.dry_run:
            conn.commit()
            print('\nCommitted.')
        else:
            print('\nDry-run complete — no changes written.')
    finally:
        conn.close()


if __name__ == '__main__':
    main()
